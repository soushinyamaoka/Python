# パッケージインストール
from selenium.webdriver.common.by import By
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl.styles import Font
from openpyxl.styles.colors import Color
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side
import time
import os
import Common.Config as Config
import pprint
import datetime
import openpyxl
from logging import getLogger
import Common.DriverCommon as DriverCommon

# loggerオブジェクトの宣言
logger = getLogger("Log")
# 定数
PURCHASE_DATE = "購入日"
AMOUNT = "金額"
PURCHASE = "購入品"
ITEM_NAME = "品名"
URL = "URL"
TITLE = "購入履歴一覧"
FILENAME = "amazon購入履歴_"
# 設定ファイル情報
conf: Config.ConfigData = None
# WebDriver
driver: webdriver.Chrome = None
# 購入履歴情報型
g_hist_info_st = {PURCHASE_DATE: "NULL", AMOUNT: "NULL", PURCHASE: "NULL"}
# 購入品情報型
g_item_info_st = {ITEM_NAME: "NULL", "URL": "NULL"}
# 全購入品履歴
g_hist_info_all = list()
# 取得年(この年からさかのぼる)
g_year = 2021

# メイン処理
# param :
#       :
# return : 無し
def main(paramConf: Config.ConfigData, paramDriver: webdriver.Chrome):
    logger.info("購入履歴EXCEL出力処理開始")
    global conf
    conf = paramConf
    global driver
    driver = paramDriver

    # 注文履歴画面へ
    order_history = driver.find_elements(By.ID, 'nav-orders')
    order_history[0].click()
    time.sleep(2)
    pages_remaining = True

    while pages_remaining:
        # 金額、購入日時情報取得
        price_data_N, date_data_N = get_ValueAndDate()

        # 商品情報取得
        item_data_N = get_ItemInfo()

        # 情報集約
        aggregation_data(price_data_N, date_data_N, item_data_N)

        # ページ切り替え
        if not DriverCommon.switchWindow(driver):
            # 切り替え不可の場合は終了
            break

    # エクセルへ出力
    output_func()
    logger.info("購入履歴EXCEL出力処理終了")

# 金額、購入日取得処理
# param : 無し
# return :
#        :
def get_ValueAndDate():
    get_element = driver.find_elements(
        By.XPATH, "//div[@class='a-row a-size-base']")
    data = [x.text for x in get_element]
    price_data = [s for s in data if "￥" in s]
    date_data = [s for s in data if "年" in s and "月" in s and "日" in s]
    price_data_N = []
    date_data_N = []
    # 金額取得
    for data in price_data:
        after_data = data.replace("￥ ", "")
        next_after_data = after_data.replace(",", "")
        price_data_N.append(next_after_data)
    # 購入日取得
    for data in date_data:
        after_data = data.replace(" ", "")
        next_after_data = after_data.replace(",", "")
        date_data_N.append(next_after_data)
    return price_data_N, date_data_N

# 商品情報取得処理
# param : 無し
# return :
def get_ItemInfo():
    item_element = driver.find_elements(
        By.XPATH, "//a[@class='a-link-normal']")
    item = [x.text for x in item_element]
    # 商品名取得
    item_data_N = list()
    for data in item:
        if "" != data and check_item_name(data):
            after_data = data.replace(" ", "")
            next_after_data = after_data.replace(",", "")
            g_item_info_st[ITEM_NAME] = next_after_data
            g_item_info_st[URL] = driver.find_elements(
                By.LINK_TEXT, data)[0].get_attribute('href')
            item_data_N.append(g_item_info_st.copy())
    return item_data_N

# 情報集約処理
# param :
#       :
#       :
# return : 無し
def aggregation_data(price_data_N, date_data_N, item_data_N):
    global g_hist_info_st
    global g_hist_info_all
    for i in range(len(price_data_N)):
        # 購入履歴情報型
        g_hist_info_st[PURCHASE_DATE] = date_data_N[i]
        g_hist_info_st[AMOUNT] = price_data_N[i]
        g_hist_info_st[PURCHASE] = item_data_N[i]
        g_hist_info_all.append(g_hist_info_st.copy())
        pprint.pprint(g_hist_info_st, width=240)

# エクセル転記処理
# param : 無し
# return : 無し
def output_func():
    global g_hist_info_all
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = TITLE
    # ヘッダ背景職
    fill = PatternFill(patternType='solid', fgColor='FFCFAD')
    #黒い実線
    side1 = Side(style='thin', color='000000')
    border_aro = Border(top=side1, bottom=side1, left=side1, right=side1)
    # 幅設定
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 90
    ws.column_dimensions["D"].width = 90
    # ヘッダ入力
    ws["A1"].value = PURCHASE_DATE
    ws["A1"].fill = fill
    ws["B1"].value = AMOUNT
    ws["B1"].fill = fill
    ws["C1"].value = ITEM_NAME
    ws["C1"].fill = fill
    ws["D1"].value = URL
    ws["D1"].fill = fill
    # 書き出し
    row_cnt = 2
    for hist_inf in g_hist_info_all:
        ws["A" + str(row_cnt)].value = hist_inf[PURCHASE_DATE]
        ws["B" + str(row_cnt)].value = int(hist_inf[AMOUNT])
        ws["B" + str(row_cnt)].number_format = openpyxl.styles.numbers.BUILTIN_FORMATS[3]
        ws["C" + str(row_cnt)].value = hist_inf[PURCHASE][ITEM_NAME]
        ws["D" + str(row_cnt)].value = hist_inf[PURCHASE][URL]
        ws["D" + str(row_cnt)].hyperlink = hist_inf[PURCHASE][URL]
        ws["D" + str(row_cnt)].font = Font(underline="single", color=Color(rgb=None, indexed=None, auto=None, theme=10, tint=0.0, type="theme"))
        row_cnt += 1
    # 罫線設定
    for row in ws["A1:D" + str(row_cnt - 1)]:
        for cell in row:
            cell.border = border_aro
    # ファイル保存
    now = datetime.datetime.now()
    file_name = FILENAME + '{}.xlsx'.format(now.strftime('%Y%m%d_%H%M%S'))
    full_path = os.path.join(conf.excel_path, file_name)
    wb.save(full_path)

# 商品名チェック処理
# param : チェック対象文字列
# return : True  商品名の場合
#        : False 商品名以外の場合
def check_item_name(name):
    #
    for text in conf.exclusion_text:
        if text == name:
            return False

    return True
